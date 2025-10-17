# ...existing code...
from flask import Flask, request, send_from_directory
from openpyxl import Workbook, load_workbook
from flask_cors import CORS
from datetime import datetime
import os

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)  # 追加

EXCEL_PATH = 'test.xlsx'

@app.route('/')
def root():
    #HTMLにアクセスしたときの最初に開くページを指定
    return send_from_directory('.', 'index.html')

@app.route('/submit', methods=['POST'])
def submit():
    
    # 固定値（ここを変更すれば送信される固定データを変えられます）
    fixed_address = '北海道千歳市美々1-2-3'
    fixed_org_name = '株式会社　サンプル商事'
    fixed_person_name = '山田　太郎'
    fixed_tel = '03-1234-5678'
    fixed_fax = '03-8765-4321'

    # フォームから受け取る項目（固定値で上書きするもの以外）
    # 日付・開始・終了を別フィールドで受け取る
    dispatch_date = request.form.get('dispatch_date')      # 例: "2025-10-17"
    dispatch_start = request.form.get('dispatch_start')    # 例: "09:00"
    dispatch_end = request.form.get('dispatch_end')        # 例: "11:00"
    # シンプルに結合（存在しない場合は空文字扱い）
    dispatch_time = f"{dispatch_date or ''} {dispatch_start or ''}〜{dispatch_end or ''}"
    dest_address = request.form.get('dest_address')
    dest_place = request.form.get('dest_place')
    dest_tel = request.form.get('dest_tel')
    dest_fax = request.form.get('dest_fax')
    reason = request.form.get('reason')

    # Excel に書き込むデータ配列（固定値を先頭に）
    data = [
        fixed_address,
        fixed_org_name,
        fixed_person_name,
        fixed_tel,
        fixed_fax,
        dispatch_time,
        dest_address,
        dest_place,
        dest_tel,
        dest_fax,
        reason
    ]

    # test.xlsxに書き込み（A列:行番号, B列:日時, C列以降:データ）
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(
            ['行番号', '書き込み日時', '住所', '名前（団体名）', '担当者名', '電話番号', 'FAX', '派遣日時', '派遣先の住所', '派遣先の会場名', '派遣先の電話番号', '派遣先のFAX', '申請事由']
        )
    else:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

    # 行番号（A列）と日時（B列）を追加
    row_number = ws.max_row  
    write_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([row_number, write_time] + data)
    wb.save(EXCEL_PATH)
    return '', 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)